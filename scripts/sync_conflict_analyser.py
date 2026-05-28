#!/usr/bin/env python3
"""
OFFICE-SYNC-CONFLICTS — deep conflict diff analyser

Answers the real question:  "What is actually different between the CONFLICT
file and the original — before I decide what to do?"

Even when the CONFLICT file is older than the original, it may contain features
or structure that the original lacks. This script shows you exactly what.

Supported file types:
  .gpkg  — per-layer feature counts + FID diff (added / removed / modified rows)
  .qgs   — semantic XML diff: relations, layouts, groups, layer sources, renderers
  .xlsx  — sheet names + row counts  (openpyxl, if available)
  other  — size + mtime comparison only

Usage:
    # Analyse ALL conflict pairs found under a folder
    python3 sync_conflict_analyser.py --scan-path /Users/g/Sync/FdI/SIG

    # Analyse a specific CONFLICT file
    python3 sync_conflict_analyser.py --file "Projeto QGIS-CONFLICT-1.qgs"
        --original "Projeto QGIS.qgs"

    # Write JSON report alongside text output
    python3 sync_conflict_analyser.py --scan-path /Users/g/Sync/FdI --json-report

    # Only show pairs where something actually differs (skip identical)
    python3 sync_conflict_analyser.py --scan-path /Users/g/Sync/FdI --only-diffs
"""

import os
import re
import sys
import json
import hashlib
import argparse
import sqlite3
import xml.etree.ElementTree as ET
from pathlib import Path
from datetime import datetime
from typing import Optional

sys.path.insert(0, str(Path(__file__).parent.parent))
try:
    from modelos.helpers import logger
    import logging
except Exception:
    import logging
    logger = logging.getLogger("sync_conflict_analyser")
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

_CONFLICT_RE = re.compile(r'^(.+?)-CONFLICT-(\d+)(\.[^/\\]+)?$', re.IGNORECASE)
_MAX_CHECKSUM_BYTES = 50 * 1024 * 1024   # 50 MB for fast checksum guard


# ──────────────────────────────────────────────────────────────────────────────
#  GeoPackage diff
# ──────────────────────────────────────────────────────────────────────────────

def _gpkg_layers(db: sqlite3.Connection) -> list:
    try:
        return [r[0] for r in db.execute(
            "SELECT table_name FROM gpkg_contents ORDER BY table_name").fetchall()]
    except Exception:
        return []


def _col_names(db: sqlite3.Connection, table: str) -> list:
    return [r[1] for r in db.execute(f'PRAGMA table_info("{table}")').fetchall()]


def _attr_hash(db: sqlite3.Connection, table: str, cols: list) -> dict:
    """fid -> MD5 of all non-geometry attribute values."""
    attr = [c for c in cols if c not in ('fid', 'geom')]
    if not attr:
        return {}
    col_expr = ', '.join(f'CAST("{c}" AS TEXT)' for c in attr)
    rows = db.execute(f'SELECT fid, {col_expr} FROM "{table}"').fetchall()
    return {r[0]: hashlib.md5('|'.join('' if v is None else v for v in r[1:]).encode()).hexdigest()
            for r in rows}


def _fid_set(db: sqlite3.Connection, table: str) -> set:
    return {r[0] for r in db.execute(f'SELECT fid FROM "{table}"').fetchall()}


def _sample_rows(db: sqlite3.Connection, table: str, fids: list, cols: list, limit=5) -> list:
    if not fids:
        return []
    attr = [c for c in cols if c not in ('geom',)]
    col_expr = ', '.join(f'"{c}"' for c in attr)
    placeholders = ','.join('?' for _ in fids[:limit])
    return db.execute(
        f'SELECT {col_expr} FROM "{table}" WHERE fid IN ({placeholders})',
        fids[:limit]).fetchall()


def diff_gpkg(conflict_path: Path, original_path: Path) -> dict:
    """Compare two GeoPackage files layer by layer. Returns structured diff."""
    result = {"type": "gpkg", "layers": {}, "error": None}
    try:
        c_db = sqlite3.connect(f"file:{conflict_path}?mode=ro", uri=True)
        o_db = sqlite3.connect(f"file:{original_path}?mode=ro", uri=True)
    except Exception as e:
        result["error"] = str(e)
        return result

    try:
        c_layers = set(_gpkg_layers(c_db))
        o_layers = set(_gpkg_layers(o_db))
        all_layers = c_layers | o_layers

        for table in sorted(all_layers):
            layer_diff = {
                "in_conflict": table in c_layers,
                "in_original": table in o_layers,
            }

            if table in c_layers and table in o_layers:
                c_cols = _col_names(c_db, table)
                o_cols = _col_names(o_db, table)
                c_fids = _fid_set(c_db, table)
                o_fids = _fid_set(o_db, table)

                only_conflict = sorted(c_fids - o_fids)
                only_original = sorted(o_fids - c_fids)
                shared = c_fids & o_fids

                c_hashes = _attr_hash(c_db, table, c_cols)
                o_hashes = _attr_hash(o_db, table, o_cols)
                modified = sorted(fid for fid in shared
                                   if c_hashes.get(fid) != o_hashes.get(fid))

                layer_diff.update({
                    "conflict_count": len(c_fids),
                    "original_count": len(o_fids),
                    "only_in_conflict": only_conflict,
                    "only_in_original": only_original,
                    "modified_fids": modified,
                    "columns_conflict": c_cols,
                    "columns_original": o_cols,
                    "columns_added_in_conflict": [c for c in c_cols if c not in o_cols],
                    "columns_removed_in_conflict": [c for c in o_cols if c not in c_cols],
                    # Sample rows for added features in CONFLICT
                    "sample_only_in_conflict": _sample_rows(c_db, table, only_conflict, c_cols),
                    "sample_only_in_original": _sample_rows(o_db, table, only_original, o_cols),
                })

                # Sample modified rows: show conflict value vs original value
                modified_samples = []
                for fid in modified[:5]:
                    attr = [c for c in c_cols if c not in ('fid', 'geom')]
                    if attr:
                        col_expr = ', '.join(f'"{c}"' for c in attr)
                        c_row = c_db.execute(f'SELECT {col_expr} FROM "{table}" WHERE fid=?', (fid,)).fetchone()
                        o_row = o_db.execute(f'SELECT {col_expr} FROM "{table}" WHERE fid=?', (fid,)).fetchone()
                        diffs = {attr[i]: {"conflict": c_row[i], "original": o_row[i]}
                                 for i in range(len(attr)) if c_row[i] != o_row[i]}
                        if diffs:
                            modified_samples.append({"fid": fid, "changes": diffs})
                layer_diff["modified_samples"] = modified_samples

            elif table in c_layers:
                fids = _fid_set(c_db, table)
                layer_diff["conflict_count"] = len(fids)
                layer_diff["note"] = "Layer exists only in CONFLICT file"
            else:
                fids = _fid_set(o_db, table)
                layer_diff["original_count"] = len(fids)
                layer_diff["note"] = "Layer exists only in ORIGINAL file"

            result["layers"][table] = layer_diff

    except Exception as e:
        result["error"] = str(e)
    finally:
        c_db.close()
        o_db.close()

    return result


# ──────────────────────────────────────────────────────────────────────────────
#  QGS diff
# ──────────────────────────────────────────────────────────────────────────────

def _qgs_parse(path: Path) -> dict:
    """Extract semantically meaningful sections from a QGS XML file."""
    info = {"save_dt": None, "save_user": None, "version": None,
            "layers": {}, "relations": {}, "layouts": [], "groups": [],
            "variables": {}, "snapping": None, "crs_description": None}
    try:
        tree = ET.parse(path)
        root = tree.getroot()
        info["save_dt"] = root.get("saveDateTime")
        info["save_user"] = root.get("saveUser")
        info["version"] = root.get("version")

        # CRS
        crs_desc = root.find(".//projectCrs//description")
        if crs_desc is not None:
            info["crs_description"] = crs_desc.text

        # Layers
        for ml in root.findall(".//maplayer"):
            lid = ml.get("id", "")
            name_el = ml.find("layername")
            src_el = ml.find("datasource")
            renderer_el = ml.find(".//renderer-v2")
            label_el = ml.find(".//labeling")
            info["layers"][lid] = {
                "name": name_el.text if name_el is not None else "",
                "source": src_el.text if src_el is not None else "",
                "renderer_type": renderer_el.get("type") if renderer_el is not None else None,
                "label_type": label_el.get("type") if label_el is not None else None,
            }

        # Relations
        for rel in root.findall(".//relation"):
            rid = rel.get("id", "")
            info["relations"][rid] = {
                "name": rel.get("name", ""),
                "ref_layer": rel.get("referencedLayerId", ""),
                "ref_field": rel.get("referencedField", ""),
                "layer": rel.get("layerId", ""),
                "field": rel.get("fieldRef", ""),
            }

        # Layouts
        info["layouts"] = [l.get("name", "") for l in root.findall(".//Layout")]

        # Layer tree groups
        info["groups"] = [g.get("name", "") for g in root.findall(".//layer-tree-group")]

        # Project variables
        for v in root.findall(".//variable"):
            info["variables"][v.get("name", "")] = v.get("value", "")

        # Snapping
        snapping_el = root.find(".//snappingConfig")
        if snapping_el is not None:
            info["snapping"] = snapping_el.attrib

    except Exception as e:
        info["parse_error"] = str(e)

    return info


def diff_qgs(conflict_path: Path, original_path: Path) -> dict:
    """Semantic diff of two QGS project files."""
    result = {"type": "qgs", "conflict_meta": {}, "original_meta": {}, "changes": {}}

    c = _qgs_parse(conflict_path)
    o = _qgs_parse(original_path)

    result["conflict_meta"] = {k: c[k] for k in ("save_dt", "save_user", "version")}
    result["original_meta"] = {k: o[k] for k in ("save_dt", "save_user", "version")}

    changes = {}

    # CRS
    if c["crs_description"] != o["crs_description"]:
        changes["crs"] = {"conflict": c["crs_description"], "original": o["crs_description"]}

    # Layers
    c_lids = set(c["layers"]); o_lids = set(o["layers"])
    layers_only_conflict = {lid: c["layers"][lid] for lid in (c_lids - o_lids)}
    layers_only_original = {lid: o["layers"][lid] for lid in (o_lids - c_lids)}
    layer_changes = {}
    for lid in c_lids & o_lids:
        cl = c["layers"][lid]; ol = o["layers"][lid]
        diffs = {}
        for field in ("source", "renderer_type", "label_type"):
            if cl[field] != ol[field]:
                diffs[field] = {"conflict": cl[field], "original": ol[field]}
        if diffs:
            layer_changes[f"{cl['name']} ({lid[:8]})"] = diffs

    if layers_only_conflict:
        changes["layers_only_in_conflict"] = {
            c["layers"][lid]["name"]: lid for lid in layers_only_conflict}
    if layers_only_original:
        changes["layers_only_in_original"] = {
            o["layers"][lid]["name"]: lid for lid in layers_only_original}
    if layer_changes:
        changes["layer_property_changes"] = layer_changes

    # Relations
    c_rids = set(c["relations"]); o_rids = set(o["relations"])
    rels_only_conflict = {rid: c["relations"][rid]["name"] for rid in (c_rids - o_rids)}
    rels_only_original = {rid: o["relations"][rid]["name"] for rid in (o_rids - c_rids)}
    if rels_only_conflict:
        changes["relations_only_in_conflict"] = rels_only_conflict
    if rels_only_original:
        changes["relations_only_in_original"] = rels_only_original

    # Layouts
    c_layouts = set(c["layouts"]); o_layouts = set(o["layouts"])
    if c_layouts != o_layouts:
        changes["layouts_only_in_conflict"] = list(c_layouts - o_layouts)
        changes["layouts_only_in_original"] = list(o_layouts - c_layouts)

    # Groups
    if c["groups"] != o["groups"]:
        changes["groups_conflict"] = c["groups"]
        changes["groups_original"] = o["groups"]

    # Variables
    c_vars = c["variables"]; o_vars = o["variables"]
    var_changes = {}
    for k in set(c_vars) | set(o_vars):
        cv = c_vars.get(k); ov = o_vars.get(k)
        if cv != ov:
            var_changes[k] = {"conflict": cv, "original": ov}
    if var_changes:
        changes["variable_changes"] = var_changes

    result["changes"] = changes
    result["has_diff"] = bool(changes)
    return result


# ──────────────────────────────────────────────────────────────────────────────
#  XLSX diff (best effort, requires openpyxl)
# ──────────────────────────────────────────────────────────────────────────────

def diff_xlsx(conflict_path: Path, original_path: Path) -> dict:
    result = {"type": "xlsx", "sheets": {}, "error": None}
    try:
        import openpyxl
        c_wb = openpyxl.load_workbook(conflict_path, read_only=True, data_only=True)
        o_wb = openpyxl.load_workbook(original_path, read_only=True, data_only=True)
        c_sheets = set(c_wb.sheetnames); o_sheets = set(o_wb.sheetnames)
        result["sheets_only_in_conflict"] = list(c_sheets - o_sheets)
        result["sheets_only_in_original"] = list(o_sheets - c_sheets)
        sheet_diffs = {}
        for name in c_sheets & o_sheets:
            c_rows = list(c_wb[name].iter_rows(values_only=True))
            o_rows = list(o_wb[name].iter_rows(values_only=True))
            c_set = set(map(str, c_rows)); o_set = set(map(str, o_rows))
            if c_set != o_set:
                sheet_diffs[name] = {
                    "conflict_rows": len(c_rows),
                    "original_rows": len(o_rows),
                    "rows_only_in_conflict": len(c_set - o_set),
                    "rows_only_in_original": len(o_set - c_set),
                }
        result["sheets"] = sheet_diffs
        result["has_diff"] = bool(sheet_diffs or result["sheets_only_in_conflict"]
                                   or result["sheets_only_in_original"])
    except ImportError:
        result["error"] = "openpyxl not installed — run: pip install openpyxl"
        result["has_diff"] = None
    except Exception as e:
        result["error"] = str(e)
        result["has_diff"] = None
    return result


# ──────────────────────────────────────────────────────────────────────────────
#  Dispatcher
# ──────────────────────────────────────────────────────────────────────────────

def deep_diff(conflict_path: Path, original_path: Path) -> dict:
    ext = conflict_path.suffix.lower()
    if ext == ".gpkg":
        return diff_gpkg(conflict_path, original_path)
    elif ext == ".qgs":
        return diff_qgs(conflict_path, original_path)
    elif ext == ".xlsx":
        return diff_xlsx(conflict_path, original_path)
    else:
        # Generic: size + mtime
        cs = conflict_path.stat(); os_ = original_path.stat()
        return {
            "type": "generic",
            "conflict_size": cs.st_size,
            "original_size": os_.st_size,
            "conflict_mtime": datetime.fromtimestamp(cs.st_mtime).isoformat(),
            "original_mtime": datetime.fromtimestamp(os_.st_mtime).isoformat(),
            "has_diff": cs.st_size != os_.st_size,
        }


# ──────────────────────────────────────────────────────────────────────────────
#  Discovery (same pattern as resolver)
# ──────────────────────────────────────────────────────────────────────────────

def discover_pairs(scan_path: Path) -> list:
    pairs = []
    for p in sorted(scan_path.rglob("*")):
        if not p.is_file():
            continue
        m = _CONFLICT_RE.match(p.name)
        if not m:
            continue
        stem, ext = m.group(1), (m.group(3) or "")
        original = p.parent / (stem + ext)
        pairs.append((p, original if original.exists() else None))
    return pairs


# ──────────────────────────────────────────────────────────────────────────────
#  Timestamp helpers
# ──────────────────────────────────────────────────────────────────────────────

def _qgs_save_dt(path: Path) -> Optional[str]:
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            for line in f:
                s = line.strip()
                if s.startswith("<qgis "):
                    m = re.search(r'saveDateTime="([^"]+)"', s)
                    return m.group(1) if m else None
                if s and not s.startswith("<?"):
                    break
    except Exception:
        pass
    return None


def _effective_ts(path: Path) -> float:
    if path.suffix.lower() == ".qgs":
        dt = _qgs_save_dt(path)
        if dt:
            try:
                return datetime.fromisoformat(dt).timestamp()
            except ValueError:
                pass
    return path.stat().st_mtime


def _ts_label(path: Path) -> str:
    if path.suffix.lower() == ".qgs":
        dt = _qgs_save_dt(path)
        if dt:
            return dt
    return datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%dT%H:%M:%S")


# ──────────────────────────────────────────────────────────────────────────────
#  Report rendering
# ──────────────────────────────────────────────────────────────────────────────

def render_report(conflict_path: Path, original_path: Optional[Path], diff: Optional[dict]) -> str:
    lines = []
    w = 70

    # Header
    lines.append("=" * w)
    try:
        rel = conflict_path.relative_to(conflict_path.parent.parent.parent.parent.parent)
    except Exception:
        rel = conflict_path
    lines.append(f"  CONFLICT: {rel}")

    if original_path is None:
        lines.append("  ⚠️  ORPHAN — original file not found")
        lines.append("=" * w)
        return "\n".join(lines)

    # Timestamp comparison
    c_ts = _effective_ts(conflict_path)
    o_ts = _effective_ts(original_path)
    c_label = _ts_label(conflict_path)
    o_label = _ts_label(original_path)
    delta_h = (o_ts - c_ts) / 3600
    verdict = "ORIGINAL is newer" if delta_h >= 0 else "⚠️  CONFLICT is NEWER"
    lines.append(f"  CONFLICT saved : {c_label}")
    lines.append(f"  Original saved : {o_label}  ← {verdict} ({abs(delta_h):.1f}h)")
    lines.append(f"  File type      : {conflict_path.suffix.lower()}")
    lines.append("-" * w)

    if diff is None:
        lines.append("  (no deep analysis available)")
        lines.append("=" * w)
        return "\n".join(lines)

    ftype = diff.get("type", "")

    # ── GeoPackage ──
    if ftype == "gpkg":
        if diff.get("error"):
            lines.append(f"  ERROR: {diff['error']}")
        else:
            any_diff = False
            for layer, ld in diff["layers"].items():
                c_cnt = ld.get("conflict_count", 0)
                o_cnt = ld.get("original_count", 0)
                only_c = ld.get("only_in_conflict", [])
                only_o = ld.get("only_in_original", [])
                modified = ld.get("modified_fids", [])
                col_added = ld.get("columns_added_in_conflict", [])
                col_removed = ld.get("columns_removed_in_conflict", [])

                layer_has_diff = bool(only_c or only_o or modified or col_added or col_removed
                                      or not ld.get("in_conflict") or not ld.get("in_original"))
                any_diff = any_diff or layer_has_diff

                if not ld.get("in_conflict"):
                    lines.append(f"  🔴 Layer '{layer}': EXISTS ONLY IN ORIGINAL ({o_cnt} features)")
                elif not ld.get("in_original"):
                    lines.append(f"  🔵 Layer '{layer}': EXISTS ONLY IN CONFLICT ({c_cnt} features)")
                else:
                    status = "✅ identical" if not layer_has_diff else "⚠️  DIFFERS"
                    lines.append(f"  Layer '{layer}': {status}  "
                                 f"(conflict={c_cnt}, original={o_cnt})")
                    if only_c:
                        lines.append(f"    + {len(only_c)} features ONLY IN CONFLICT  "
                                     f"(fids: {only_c[:5]}{'...' if len(only_c)>5 else ''})")
                        for row in ld.get("sample_only_in_conflict", [])[:3]:
                            lines.append(f"      sample: {row}")
                    if only_o:
                        lines.append(f"    - {len(only_o)} features ONLY IN ORIGINAL  "
                                     f"(fids: {only_o[:5]}{'...' if len(only_o)>5 else ''})")
                        for row in ld.get("sample_only_in_original", [])[:3]:
                            lines.append(f"      sample: {row}")
                    if modified:
                        lines.append(f"    ~ {len(modified)} features with MODIFIED attributes  "
                                     f"(fids: {modified[:5]}{'...' if len(modified)>5 else ''})")
                        for ms in ld.get("modified_samples", [])[:3]:
                            lines.append(f"      fid={ms['fid']}: {ms['changes']}")
                    if col_added:
                        lines.append(f"    + columns added in CONFLICT: {col_added}")
                    if col_removed:
                        lines.append(f"    - columns removed in CONFLICT: {col_removed}")

            if not any_diff:
                lines.append("  ✅ All layers IDENTICAL — safe to delete CONFLICT")

    # ── QGS ──
    elif ftype == "qgs":
        changes = diff.get("changes", {})
        if not changes:
            lines.append("  ✅ No semantic differences found — safe to delete CONFLICT")
        else:
            lines.append(f"  ⚠️  {len(changes)} section(s) differ:")

            if "crs" in changes:
                lines.append(f"    CRS:")
                lines.append(f"      CONFLICT : {changes['crs']['conflict']}")
                lines.append(f"      Original : {changes['crs']['original']}")

            if "layers_only_in_conflict" in changes:
                lines.append(f"    Layers only in CONFLICT ({len(changes['layers_only_in_conflict'])}):")
                for name in list(changes["layers_only_in_conflict"])[:5]:
                    lines.append(f"      + {name}")

            if "layers_only_in_original" in changes:
                lines.append(f"    Layers only in ORIGINAL ({len(changes['layers_only_in_original'])}):")
                for name in list(changes["layers_only_in_original"])[:5]:
                    lines.append(f"      - {name}")

            if "layer_property_changes" in changes:
                lines.append(f"    Layer property changes ({len(changes['layer_property_changes'])}):")
                for name, diffs in list(changes["layer_property_changes"].items())[:5]:
                    for field, vals in diffs.items():
                        lines.append(f"      {name}.{field}:")
                        lines.append(f"        CONFLICT : {vals['conflict']}")
                        lines.append(f"        Original : {vals['original']}")

            if "relations_only_in_conflict" in changes:
                items = changes["relations_only_in_conflict"]
                lines.append(f"    Relations only in CONFLICT ({len(items)}):")
                for name in list(set(items.values()))[:5]:
                    lines.append(f"      + {name}")

            if "relations_only_in_original" in changes:
                items = changes["relations_only_in_original"]
                lines.append(f"    Relations only in ORIGINAL ({len(items)}):")
                for name in list(set(items.values()))[:5]:
                    lines.append(f"      - {name}")

            if "layouts_only_in_conflict" in changes and changes["layouts_only_in_conflict"]:
                lines.append(f"    Layouts only in CONFLICT: {changes['layouts_only_in_conflict']}")
            if "layouts_only_in_original" in changes and changes["layouts_only_in_original"]:
                lines.append(f"    Layouts only in ORIGINAL: {changes['layouts_only_in_original']}")

            if "variable_changes" in changes:
                lines.append(f"    Variable changes: {changes['variable_changes']}")

            if "groups_conflict" in changes:
                lines.append(f"    Layer groups differ:")
                lines.append(f"      CONFLICT : {changes['groups_conflict']}")
                lines.append(f"      Original : {changes['groups_original']}")

    # ── XLSX ──
    elif ftype == "xlsx":
        if diff.get("error"):
            lines.append(f"  NOTE: {diff['error']}")
        elif not diff.get("sheets") and not diff.get("sheets_only_in_conflict") and not diff.get("sheets_only_in_original"):
            lines.append("  ✅ All sheets identical")
        else:
            if diff.get("sheets_only_in_conflict"):
                lines.append(f"  Sheets only in CONFLICT: {diff['sheets_only_in_conflict']}")
            if diff.get("sheets_only_in_original"):
                lines.append(f"  Sheets only in ORIGINAL: {diff['sheets_only_in_original']}")
            for sheet, sd in diff.get("sheets", {}).items():
                lines.append(f"  Sheet '{sheet}': conflict={sd['conflict_rows']} rows, "
                             f"original={sd['original_rows']} rows")
                if sd.get("rows_only_in_conflict"):
                    lines.append(f"    + {sd['rows_only_in_conflict']} rows only in CONFLICT")
                if sd.get("rows_only_in_original"):
                    lines.append(f"    - {sd['rows_only_in_original']} rows only in ORIGINAL")

    # ── Generic ──
    elif ftype == "generic":
        if not diff.get("has_diff"):
            lines.append("  ✅ Same file size")
        else:
            lines.append(f"  Size differs: conflict={diff['conflict_size']} bytes, "
                         f"original={diff['original_size']} bytes")

    # Resolution hint
    lines.append("-" * w)
    has_diff = diff.get("has_diff", True)
    conflict_newer = o_ts < c_ts
    if has_diff is False and not conflict_newer:
        lines.append("  RECOMMENDATION: ✅ Safe to delete — original is newer and content is identical")
    elif has_diff is False and conflict_newer:
        lines.append("  RECOMMENDATION: ✅ Safe to delete — conflict is newer BUT content is identical")
    elif conflict_newer and has_diff:
        lines.append("  RECOMMENDATION: ⚠️  REVIEW — conflict is NEWER and has different content")
        lines.append("                  → Check diffs above; import missing features or copy QGS sections")
    elif not conflict_newer and has_diff:
        lines.append("  RECOMMENDATION: ⚠️  REVIEW — original is newer but CONFLICT has extra content")
        lines.append("                  → Verify CONFLICT features are already in original before deleting")
    else:
        lines.append("  RECOMMENDATION: ⚠️  REVIEW — manual verification needed")

    lines.append("=" * w)
    return "\n".join(lines)


# ──────────────────────────────────────────────────────────────────────────────
#  Main
# ──────────────────────────────────────────────────────────────────────────────

def main() -> int:
    parser = argparse.ArgumentParser(
        description="Deep diff analyser for sync.com CONFLICT file pairs",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--scan-path",
                        default=os.getenv("SIG_BASE", "/Users/g/Sync/FdI/SIG"),
                        help="Folder to scan recursively")
    parser.add_argument("--file", help="Specific CONFLICT file path")
    parser.add_argument("--original", help="Matching original (required with --file)")
    parser.add_argument("--json-report", action="store_true",
                        help="Write JSON report to logs/")
    parser.add_argument("--only-diffs", action="store_true",
                        help="Skip pairs that are identical")
    parser.add_argument("--verbose", action="store_true")
    args = parser.parse_args()

    if args.file:
        conflict = Path(args.file)
        original = Path(args.original) if args.original else None
        pairs = [(conflict, original)]
    else:
        pairs = discover_pairs(Path(args.scan_path))

    if not pairs:
        print("No CONFLICT files found.")
        return 0

    all_results = []
    total = len(pairs)
    print(f"\nAnalysing {total} CONFLICT file(s) under {args.scan_path if not args.file else args.file}\n")

    for i, (conflict_path, original_path) in enumerate(pairs, 1):
        ext = conflict_path.suffix.lower()
        # Skip unsupported large binaries that aren't interesting to diff
        if ext in (".tif", ".tiff", ".png", ".jpg", ".jpeg", ".pdf"):
            continue
        if ext == ".qgs~":
            continue  # backup files, skip

        logger.info(f"[{i}/{total}] {conflict_path.name}")

        diff_result = None
        if original_path is not None:
            try:
                diff_result = deep_diff(conflict_path, original_path)
            except Exception as e:
                diff_result = {"type": "error", "error": str(e), "has_diff": None}

        if args.only_diffs and diff_result and diff_result.get("has_diff") is False:
            continue

        report_text = render_report(conflict_path, original_path, diff_result)
        print(report_text)

        all_results.append({
            "conflict": str(conflict_path),
            "original": str(original_path) if original_path else None,
            "diff": diff_result,
        })

    if args.json_report and all_results:
        log_dir = Path(os.getenv("LOG_DIR",
                       str(Path(__file__).parent.parent / "logs")))
        log_dir.mkdir(parents=True, exist_ok=True)
        out = log_dir / f"sync_conflict_analysis_{datetime.now():%Y-%m-%d}.json"
        with open(out, "w") as f:
            json.dump(all_results, f, indent=2, default=str)
        print(f"JSON report written to {out}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
